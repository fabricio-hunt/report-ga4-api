"""Excel export with openpyxl formatting."""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.io.paths import resolve_output_dir

logger = logging.getLogger("ga4")

FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
FONT_HEADER = Font(name="Arial", color="FFFFFF", bold=True, size=11)
FONT_DATA = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=13, color="1F4E79")

THIN = Side(style="thin", color="BFBFBF")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_KEYWORDS = ["Receita"]
PCT_KEYWORDS = ["Taxa de engajamento", "Engajamento"]


def _col_number_format(col_name: str) -> Optional[str]:
    if any(keyword in col_name for keyword in CURRENCY_KEYWORDS):
        return "R$ #,##0.00"
    if any(keyword in col_name for keyword in PCT_KEYWORDS):
        return '0.00"%"'
    return None


def _format_sheet(ws, title: str, period: str) -> None:
    ws.insert_rows(1, 2)
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A2"] = f"Período: {period}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="595959")

    header_row = 3
    data_start = 4

    for cell in ws[header_row]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    ws.row_dimensions[header_row].height = 30
    max_row = ws.max_row

    for row_idx in range(data_start, max_row + 1):
        fill = FILL_ROW_ALT if (row_idx - data_start) % 2 == 1 else None
        for cell in ws[row_idx]:
            cell.font = FONT_DATA
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

            if cell.column > 2:
                col_name = ws.cell(row=header_row, column=cell.column).value or ""
                fmt = _col_number_format(col_name)
                if fmt:
                    cell.number_format = fmt
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

    for col in ws.iter_cols(min_row=header_row, max_row=max_row):
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 32)

    ws.freeze_panes = ws.cell(row=data_start, column=1)


def export_sheets_to_excel(
    sheets: dict[str, tuple[str, str, pd.DataFrame]],
    filename: str,
    output_dir: str | None = None,
) -> str:
    """Write formatted Excel file from sheet definitions."""
    if not sheets:
        logger.warning("Nenhum dado coletado para exportar.")
        return ""

    try:
        target_dir = output_dir or resolve_output_dir()
        filepath = os.path.join(target_dir, filename)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for sheet_name, (_, _, df) in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(filepath)
        for sheet_name, (title, period, _) in sheets.items():
            _format_sheet(wb[sheet_name], title=title, period=period)
        wb.save(filepath)

        logger.info("Arquivo salvo: %s", filepath)
        return filepath
    except Exception as exc:
        logger.error("Erro na exportação: %s", exc)
        return ""


def export_main_report(
    df_web: Optional[pd.DataFrame],
    df_farma: Optional[pd.DataFrame],
    df_app: Optional[pd.DataFrame],
    period_label: str,
    analysis_start: str,
    output_dir: str | None = None,
) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"GA4_Bemol_{analysis_start[:7]}_{timestamp}.xlsx"

    sheets: dict[str, tuple[str, str, pd.DataFrame]] = {}
    if df_web is not None and not df_web.empty:
        sheets["Bemol Web"] = ("Bemol Web — Orgânico + Total", period_label, df_web)
    if df_farma is not None and not df_farma.empty:
        sheets["Bemol Farma"] = ("Bemol Farma — Orgânico", period_label, df_farma)
    if df_app is not None and not df_app.empty:
        sheets["Bemol App"] = (
            "Bemol App — Orgânico + Total (Android & iOS)",
            period_label,
            df_app,
        )

    return export_sheets_to_excel(sheets, filename, output_dir)


def export_bemol_app_report(
    df_farma_web: Optional[pd.DataFrame],
    period_label: str,
    analysis_start: str,
    analysis_end: str,
    output_dir: str | None = None,
) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = (
        f"GA4_Farma_Web_Organic_{analysis_start[:4]}-{analysis_end[:4]}_{timestamp}.xlsx"
    )

    sheets: dict[str, tuple[str, str, pd.DataFrame]] = {}
    if df_farma_web is not None and not df_farma_web.empty:
        sheets["Bemol Farma Web"] = ("Farma (Web + Orgânico)", period_label, df_farma_web)

    return export_sheets_to_excel(sheets, filename, output_dir)


def export_farma_comparacao_report(
    df_farma: Optional[pd.DataFrame],
    period_label: str,
    output_dir: str | None = None,
) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"GA4_Report_Farma_Total_{timestamp}.xlsx"

    sheets: dict[str, tuple[str, str, pd.DataFrame]] = {}
    if df_farma is not None and not df_farma.empty:
        sheets["Bemol Farma Web Total"] = (
            "Farma — Web Total (Sessões, Usuários, Engajamento, Receita, Transações)",
            period_label,
            df_farma,
        )

    return export_sheets_to_excel(sheets, filename, output_dir)
