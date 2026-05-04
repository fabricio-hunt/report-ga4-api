"""
GA4 Data Collector - Bemol Group (VERSÃO 4.0 — Comparação YoY)
Coleta dados orgânicos com comparação mês a mês entre 2025 e 2026.

ATUALIZAÇÕES v4.0:
1. Comparação YoY nativa via múltiplos date_ranges na mesma requisição GA4.
2. Colunas separadas por sufixo _2025 / _2026 + coluna delta_pct por métrica.
3. Config.COMPARISON_OFFSET: calcula automaticamente o período equivalente 2025.
4. Processamento centralizado em _process_yoy_rows() — sem duplicação de lógica.
5. Excel com formatação condicional: delta positivo = verde, negativo = vermelho.

Author: Analytics Team
Version: 4.0
"""

import os
import logging
import pickle
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import (
    RunReportRequest, DateRange, Dimension, Metric,
    FilterExpression, Filter, FilterExpressionList
)

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

class Config:
    """Configurações centralizadas."""

    PROPERTIES = {
        'ecommerce_bemol': '272846783',
        'bemol_farma': '374507450',
    }

    CLIENT_SECRET_FILE = 'client_secret.json'
    TOKEN_FILE = 'token.pickle'
    SCOPES = ['https://www.googleapis.com/auth/analytics.readonly']

    # --- Período de análise (2026) ---
    ANALYSIS_START = '2026-03-01'
    ANALYSIS_END   = '2026-03-31'

    # --- Período de comparação (calculado automaticamente: -1 ano) ---
    @classmethod
    def comparison_period(cls) -> tuple[str, str]:
        """Retorna (start, end) do mesmo período no ano anterior."""
        start = date.fromisoformat(cls.ANALYSIS_START) - relativedelta(years=1)
        end   = date.fromisoformat(cls.ANALYSIS_END)   - relativedelta(years=1)
        return start.isoformat(), end.isoformat()

    OUTPUT_DIR = 'ga4_reports'
    LOG_FILE   = 'ga4_yoy_collector.log'

    ORGANIC_SOURCES = [
        'google-play / organic',
        'google / organic',
        'bing / organic',
        'yahoo / organic',
        'duckduckgo / organic',
        'ecosia.org / organic',
        'yandex / organic',
        'awin / organic',
    ]

# ============================================================================
# HELPERS — FILTROS
# ============================================================================

MONTH_NAMES = {
    '01': 'janeiro', '02': 'fevereiro', '03': 'março',
    '04': 'abril',   '05': 'maio',      '06': 'junho',
    '07': 'julho',   '08': 'agosto',    '09': 'setembro',
    '10': 'outubro', '11': 'novembro',  '12': 'dezembro',
}


def _or_filter(field: str, values: list[str]) -> FilterExpression:
    return FilterExpression(
        or_group=FilterExpressionList(expressions=[
            FilterExpression(filter=Filter(
                field_name=field,
                string_filter=Filter.StringFilter(
                    match_type=Filter.StringFilter.MatchType.EXACT,
                    value=v
                )
            )) for v in values
        ])
    )


def _and_filter(*expressions: FilterExpression) -> FilterExpression:
    return FilterExpression(
        and_group=FilterExpressionList(expressions=list(expressions))
    )


def organic_filter() -> FilterExpression:
    return _or_filter('sessionSourceMedium', Config.ORGANIC_SOURCES)


def app_platform_filter() -> FilterExpression:
    return _or_filter('platform', ['Android', 'iOS'])


def web_platform_filter() -> FilterExpression:
    return _or_filter('platform', ['web'])

# ============================================================================
# LOGGING & AUTENTICAÇÃO
# ============================================================================

def setup_logging() -> logging.Logger:
    os.makedirs(Config.OUTPUT_DIR, exist_ok=True)
    log_path = os.path.join(Config.OUTPUT_DIR, Config.LOG_FILE)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(levelname)-8s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        handlers=[
            logging.FileHandler(log_path, encoding='utf-8'),
            logging.StreamHandler(),
        ]
    )
    return logging.getLogger(__name__)


logger = setup_logging()


def authenticate_ga4() -> Optional[BetaAnalyticsDataClient]:
    try:
        creds = None
        if os.path.exists(Config.TOKEN_FILE):
            with open(Config.TOKEN_FILE, 'rb') as f:
                creds = pickle.load(f)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    Config.CLIENT_SECRET_FILE, Config.SCOPES
                )
                creds = flow.run_local_server(port=0)
            with open(Config.TOKEN_FILE, 'wb') as f:
                pickle.dump(creds, f)
        return BetaAnalyticsDataClient(credentials=creds)
    except Exception as e:
        logger.error(f"Erro na autenticação: {e}")
        return None

# ============================================================================
# PROCESSAMENTO YoY — NÚCLEO DA v4.0
# ============================================================================

def _date_ranges() -> list[DateRange]:
    """
    Retorna dois DateRange: [0] = 2026 (atual), [1] = 2025 (comparação).

    A API GA4 rotula as linhas com a dimensão 'dateRange' contendo
    'date_range_0' e 'date_range_1' quando múltiplos intervalos são passados.
    """
    comp_start, comp_end = Config.comparison_period()
    return [
        DateRange(start_date=Config.ANALYSIS_START, end_date=Config.ANALYSIS_END),
        DateRange(start_date=comp_start, end_date=comp_end),
    ]


def _process_yoy_rows(
    rows,
    metric_map: dict[int, str],
    extra_dimensions: list[str] | None = None,
) -> pd.DataFrame:
    """
    Converte as linhas da resposta GA4 (com múltiplos date_ranges) em
    um DataFrame com colunas sufixadas _2026 / _2025 + delta_pct.

    Args:
        rows: response.rows da API GA4.
        metric_map: {índice_posição: nome_coluna} — ex: {0: 'Sessões orgânicas'}.
        extra_dimensions: nomes adicionais após (month, year) para agrupar.

    A dimensão 'dateRange' sempre ocupa a ÚLTIMA posição nas dimensões
    quando múltiplos date_ranges são usados.
    """
    buckets: dict[str, dict] = {}

    for row in rows:
        dims = [d.value for d in row.dimension_values]

        # Última dimensão = dateRange ('date_range_0' ou 'date_range_1')
        date_range_label = dims[-1]
        month = dims[0]
        year  = dims[1]
        extra = tuple(dims[2:-1]) if extra_dimensions else ()

        key = (month, year) + extra
        suffix = '_2026' if date_range_label == 'date_range_0' else '_2025'

        if key not in buckets:
            buckets[key] = {}

        for pos, col_name in metric_map.items():
            raw_val = row.metric_values[pos].value
            # Detecta float vs int pelo ponto decimal
            value = float(raw_val) if '.' in raw_val else int(raw_val)
            buckets[key][col_name + suffix] = value

    if not buckets:
        return pd.DataFrame()

    records = []
    for key, metrics in sorted(buckets.items()):
        month, year = key[0], key[1]
        extra = key[2:]
        record: dict = {
            'Mês': MONTH_NAMES.get(month, month),
            'Ano base': year,
        }
        if extra_dimensions:
            for name, val in zip(extra_dimensions, extra):
                record[name] = val

        record.update(metrics)

        # Delta percentual por métrica (evita divisão por zero)
        for col_name in metric_map.values():
            v26 = metrics.get(col_name + '_2026', 0)
            v25 = metrics.get(col_name + '_2025', 0)
            if v25 and v25 != 0:
                record[col_name + '_delta_pct'] = round((v26 - v25) / v25 * 100, 2)
            else:
                record[col_name + '_delta_pct'] = None

        records.append(record)

    return pd.DataFrame(records)

# ============================================================================
# COLETA — BEMOL WEB
# ============================================================================

def fetch_bemol_web(client: BetaAnalyticsDataClient) -> Optional[pd.DataFrame]:
    logger.info("🌐 Coletando: Bemol Web (orgânico) — YoY")
    prop = f"properties/{Config.PROPERTIES['ecommerce_bemol']}"

    try:
        # --- Orgânico ---
        req_org = RunReportRequest(
            property=prop,
            dimensions=[
                Dimension(name='month'),
                Dimension(name='year'),
                Dimension(name='dateRange'),   # obrigatório para YoY
            ],
            metrics=[
                Metric(name='sessions'),
                Metric(name='activeUsers'),
                Metric(name='engagementRate'),
                Metric(name='totalRevenue'),
            ],
            date_ranges=_date_ranges(),
            dimension_filter=_and_filter(web_platform_filter(), organic_filter()),
        )
        resp_org = client.run_report(request=req_org)

        df_org = _process_yoy_rows(
            resp_org.rows,
            metric_map={
                0: 'Sessões orgânicas',
                1: 'Usuários orgânicos',
                2: 'Taxa de engajamento (%)',
                3: 'Receita orgânica',
            },
        )
        # Taxa de engajamento vem como 0-1, converte para %
        for suffix in ('_2026', '_2025'):
            col = f'Taxa de engajamento (%)' + suffix
            if col in df_org.columns:
                df_org[col] = (df_org[col] * 100).round(2)

        # --- Total (todos os canais) ---
        req_tot = RunReportRequest(
            property=prop,
            dimensions=[
                Dimension(name='month'),
                Dimension(name='year'),
                Dimension(name='dateRange'),
            ],
            metrics=[
                Metric(name='sessions'),
                Metric(name='totalRevenue'),
            ],
            date_ranges=_date_ranges(),
            dimension_filter=web_platform_filter(),
        )
        resp_tot = client.run_report(request=req_tot)

        df_tot = _process_yoy_rows(
            resp_tot.rows,
            metric_map={
                0: 'Sessões totais',
                1: 'Receita total',
            },
        )

        df = df_org.merge(df_tot, on=['Mês', 'Ano base'], how='outer')
        return df

    except Exception as e:
        logger.error(f"Erro Bemol Web: {e}")
        return None

# ============================================================================
# COLETA — BEMOL FARMA
# ============================================================================

def fetch_bemol_farma(client: BetaAnalyticsDataClient) -> Optional[pd.DataFrame]:
    logger.info("🏥 Coletando: Bemol Farma (orgânico) — YoY")
    prop = f"properties/{Config.PROPERTIES['bemol_farma']}"

    try:
        req = RunReportRequest(
            property=prop,
            dimensions=[
                Dimension(name='month'),
                Dimension(name='year'),
                Dimension(name='dateRange'),
            ],
            metrics=[
                Metric(name='sessions'),
                Metric(name='activeUsers'),
                Metric(name='engagementRate'),
                Metric(name='totalRevenue'),
            ],
            date_ranges=_date_ranges(),
            dimension_filter=organic_filter(),
        )
        resp = client.run_report(request=req)

        df = _process_yoy_rows(
            resp.rows,
            metric_map={
                0: 'Sessões orgânicas',
                1: 'Usuários orgânicos',
                2: 'Taxa de engajamento (%)',
                3: 'Receita orgânica',
            },
        )
        for suffix in ('_2026', '_2025'):
            col = 'Taxa de engajamento (%)' + suffix
            if col in df.columns:
                df[col] = (df[col] * 100).round(2)

        return df

    except Exception as e:
        logger.error(f"Erro Bemol Farma: {e}")
        return None

# ============================================================================
# COLETA — BEMOL APP
# ============================================================================

def fetch_bemol_app(client: BetaAnalyticsDataClient) -> Optional[pd.DataFrame]:
    logger.info("📱 Coletando: Bemol App (orgânico) — YoY")
    prop = f"properties/{Config.PROPERTIES['ecommerce_bemol']}"

    try:
        # --- Orgânico ---
        req_org = RunReportRequest(
            property=prop,
            dimensions=[
                Dimension(name='month'),
                Dimension(name='year'),
                Dimension(name='dateRange'),
            ],
            metrics=[
                Metric(name='sessions'),
                Metric(name='activeUsers'),
                Metric(name='newUsers'),
                Metric(name='transactions'),
                Metric(name='totalRevenue'),
            ],
            date_ranges=_date_ranges(),
            dimension_filter=_and_filter(app_platform_filter(), organic_filter()),
        )
        resp_org = client.run_report(request=req_org)

        df_org = _process_yoy_rows(
            resp_org.rows,
            metric_map={
                0: 'Sessões organic',
                1: 'Usuários ativos organic',
                2: 'Novos usuários organic',
                3: 'Transações',
                4: 'Receita organic',
            },
        )

        # --- Total ---
        req_tot = RunReportRequest(
            property=prop,
            dimensions=[
                Dimension(name='month'),
                Dimension(name='year'),
                Dimension(name='dateRange'),
            ],
            metrics=[
                Metric(name='activeUsers'),
                Metric(name='sessions'),
                Metric(name='totalRevenue'),
            ],
            date_ranges=_date_ranges(),
            dimension_filter=app_platform_filter(),
        )
        resp_tot = client.run_report(request=req_tot)

        df_tot = _process_yoy_rows(
            resp_tot.rows,
            metric_map={
                0: 'App Usuários ativos total',
                1: 'App Sessões total',
                2: 'App Receita total',
            },
        )

        df = df_org.merge(df_tot, on=['Mês', 'Ano base'], how='outer')
        return df

    except Exception as e:
        logger.error(f"Erro Bemol App: {e}")
        return None

# ============================================================================
# EXPORTAÇÃO EXCEL COM FORMATAÇÃO CONDICIONAL
# ============================================================================

FILL_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_RED   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_HEADER = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FONT_HEADER = Font(color='FFFFFF', bold=True)


def _format_sheet(ws) -> None:
    """Aplica formatação básica e cores nas colunas delta_pct."""
    # Cabeçalho
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    # Largura automática
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            max_len + 4, 14
        )

    # Pintura condicional nas colunas delta_pct
    delta_cols = [
        cell.column for cell in ws[1]
        if cell.value and '_delta_pct' in str(cell.value)
    ]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in delta_cols and cell.value is not None:
                cell.fill = FILL_GREEN if cell.value >= 0 else FILL_RED


def export_to_excel(
    df_web: Optional[pd.DataFrame],
    df_farma: Optional[pd.DataFrame],
    df_app: Optional[pd.DataFrame],
) -> str:
    try:
        comp_start, _ = Config.comparison_period()
        comp_year = comp_start[:4]
        curr_year = Config.ANALYSIS_START[:4]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'GA4_Bemol_YoY_{comp_year}vs{curr_year}_{timestamp}.xlsx'
        filepath = os.path.join(Config.OUTPUT_DIR, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            if df_web is not None:
                df_web.to_excel(writer, sheet_name='Bemol Web', index=False)
            if df_farma is not None:
                df_farma.to_excel(writer, sheet_name='Bemol Farma', index=False)
            if df_app is not None:
                df_app.to_excel(writer, sheet_name='Bemol App', index=False)

        wb = load_workbook(filepath)
        for sheet_name in wb.sheetnames:
            _format_sheet(wb[sheet_name])
        wb.save(filepath)

        logger.info(f"✅ Arquivo salvo: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Erro exportação: {e}")
        return ''

# ============================================================================
# MAIN
# ============================================================================

def main() -> None:
    comp_start, comp_end = Config.comparison_period()
    print(f"""
GA4 YoY COLLECTOR v4.0
  Período atual:     {Config.ANALYSIS_START} → {Config.ANALYSIS_END}
  Período anterior:  {comp_start} → {comp_end}
""")

    client = authenticate_ga4()
    if not client:
        return

    df_web   = fetch_bemol_web(client)
    df_farma = fetch_bemol_farma(client)
    df_app   = fetch_bemol_app(client)

    path = export_to_excel(df_web, df_farma, df_app)
    if path:
        print(f"\n✅ Relatório YoY gerado: {path}")


if __name__ == '__main__':
    main()