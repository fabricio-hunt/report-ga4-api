"""
GA4 Data Collector - Bemol Group
Coleta dados completos para preenchimento das planilhas:
- Bemol (web): Sessões e Receita totais
- Bemol (App): Usuários ativos, Sessões e Receita
- Bemol Farma: Sessões orgânicas, Usuários orgânicos, Taxa engajamento, Receita orgânica

Author: Analytics Team
Version: 2.0
"""

import os
import logging
import pickle
from datetime import datetime
from typing import Optional, Dict, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import (
    RunReportRequest, DateRange, Dimension, Metric, 
    OrderBy, FilterExpression, Filter
)

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

class Config:
    """Configurações centralizadas"""
    # IDs das propriedades GA4
    PROPERTIES = {
        'ecommerce_bemol': '272846783',  # Ecommerce Bemol (desk, mobile e APP)
        'bemol_farma': '374507450'        # Bemol Farma - GA4
    }
    
    CLIENT_SECRET_FILE = 'client_secret.json'
    TOKEN_FILE = 'token.pickle'
    SCOPES = ['https://www.googleapis.com/auth/analytics.readonly']
    
    # Períodos de análise
    ANALYSIS_YEAR = '2026'
    ANALYSIS_START = '2026-01-01'
    ANALYSIS_END = '2026-12-31'
    
    # Output
    OUTPUT_DIR = 'ga4_reports'
    LOG_FILE = 'ga4_complete_collector.log'

# ============================================================================
# LOGGING
# ============================================================================

def setup_logging() -> logging.Logger:
    """Configura logging"""
    os.makedirs(Config.OUTPUT_DIR, exist_ok=True)
    log_path = os.path.join(Config.OUTPUT_DIR, Config.LOG_FILE)
    
    log_format = '%(asctime)s | %(levelname)-8s | %(message)s'
    date_format = '%Y-%m-%d %H:%M:%S'
    
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        datefmt=date_format,
        handlers=[
            logging.FileHandler(log_path, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info("="*80)
    logger.info("GA4 Complete Data Collector - Bemol Group")
    logger.info("="*80)
    
    return logger

logger = setup_logging()

# ============================================================================
# AUTENTICAÇÃO
# ============================================================================

def authenticate_ga4() -> Optional[BetaAnalyticsDataClient]:
    """Autentica no Google Analytics 4"""
    try:
        logger.info("Iniciando autenticação")
        
        credentials = None
        
        if os.path.exists(Config.TOKEN_FILE):
            with open(Config.TOKEN_FILE, 'rb') as token:
                credentials = pickle.load(token)
            logger.info("✓ Credenciais carregadas")
        
        if credentials and credentials.valid:
            logger.info("✓ Credenciais válidas")
        elif credentials and credentials.expired and credentials.refresh_token:
            logger.info("Renovando credenciais...")
            try:
                credentials.refresh(Request())
                logger.info("✓ Credenciais renovadas")
            except Exception as e:
                logger.warning(f"Falha ao renovar: {e}")
                credentials = None
        else:
            credentials = None
        
        if not credentials:
            logger.info("Nova autenticação necessária")
            
            if not os.path.exists(Config.CLIENT_SECRET_FILE):
                logger.error(f"Arquivo não encontrado: {Config.CLIENT_SECRET_FILE}")
                return None
            
            flow = InstalledAppFlow.from_client_secrets_file(
                Config.CLIENT_SECRET_FILE, 
                Config.SCOPES
            )
            
            print("\n" + "="*80)
            print("🔐 AUTENTICAÇÃO OAUTH 2.0")
            print("="*80)
            print("Uma janela do navegador será aberta.")
            print("Faça login e autorize o acesso ao Google Analytics.")
            print("="*80 + "\n")
            
            try:
                credentials = flow.run_local_server(
                    port=0,
                    success_message='✅ Autenticação concluída! Você pode fechar esta janela.',
                    open_browser=True
                )
            except KeyboardInterrupt:
                logger.error("Autenticação cancelada")
                return None
            
            with open(Config.TOKEN_FILE, 'wb') as token:
                pickle.dump(credentials, token)
            logger.info("✓ Credenciais salvas")
        
        client = BetaAnalyticsDataClient(credentials=credentials)
        logger.info("✓ Cliente GA4 criado com sucesso")
        
        return client
        
    except Exception as e:
        logger.error(f"Erro na autenticação: {str(e)}", exc_info=True)
        return None

# ============================================================================
# FUNÇÕES DE COLETA - BEMOL WEB
# ============================================================================

def fetch_bemol_web_data(
    client: BetaAnalyticsDataClient,
    property_id: str,
    start_date: str,
    end_date: str
) -> Optional[pd.DataFrame]:
    """
    Coleta dados para Bemol (web):
    - Sessões totais (todos os canais)
    - Receita total (todos os canais)
    Filtro: platform = 'web'
    """
    try:
        logger.info("📊 Coletando dados: Bemol (web) - Sessões e Receita totais")
        
        # Filtro de plataforma web
        platform_filter = FilterExpression(
            filter=Filter(
                field_name='platform',
                string_filter=Filter.StringFilter(
                    match_type=Filter.StringFilter.MatchType.EXACT,
                    value='web'
                )
            )
        )
        
        request = RunReportRequest(
            property=f"properties/{property_id}",
            dimensions=[
                Dimension(name="month"),
                Dimension(name="year")
            ],
            metrics=[
                Metric(name="sessions"),
                Metric(name="totalRevenue")
            ],
            date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
            dimension_filter=platform_filter,
            order_bys=[
                OrderBy(dimension={'dimension_name': 'year'}),
                OrderBy(dimension={'dimension_name': 'month'})
            ]
        )
        
        response = client.run_report(request=request)
        
        data = []
        for row in response.rows:
            month = row.dimension_values[0].value
            year = row.dimension_values[1].value
            
            data.append({
                'Mês': month,
                'Ano': year,
                'Sessões totais (todos os canais)': int(row.metric_values[0].value),
                'Receita total (todos os canais)': float(row.metric_values[1].value)
            })
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            # Converter números de mês para nomes em português
            month_names = {
                '01': 'janeiro', '02': 'fevereiro', '03': 'março',
                '04': 'abril', '05': 'maio', '06': 'junho',
                '07': 'julho', '08': 'agosto', '09': 'setembro',
                '10': 'outubro', '11': 'novembro', '12': 'dezembro'
            }
            df['Mês'] = df['Mês'].map(month_names)
            
            logger.info(f"✓ Coletados {len(df)} meses")
            logger.info(f"  Total Sessões: {df['Sessões totais (todos os canais)'].sum():,.0f}".replace(',', '.'))
            logger.info(f"  Total Receita: R$ {df['Receita total (todos os canais)'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        return df
        
    except Exception as e:
        logger.error(f"Erro ao coletar Bemol Web: {str(e)}", exc_info=True)
        return None

# ============================================================================
# FUNÇÕES DE COLETA - BEMOL APP
# ============================================================================

def fetch_bemol_app_data(
    client: BetaAnalyticsDataClient,
    property_id: str,
    start_date: str,
    end_date: str
) -> Optional[pd.DataFrame]:
    """
    Coleta dados para Bemol (App):
    - Usuários ativos
    - Sessões
    - Receita
    Filtros: platform IN ('Android', 'iOS')
    """
    try:
        logger.info("📱 Coletando dados: Bemol (App) - Usuários, Sessões e Receita")
        
        # Filtro para Android ou iOS
        platform_filter = FilterExpression(
            or_group=FilterExpression.FilterExpressionList(
                expressions=[
                    FilterExpression(
                        filter=Filter(
                            field_name='platform',
                            string_filter=Filter.StringFilter(
                                match_type=Filter.StringFilter.MatchType.EXACT,
                                value='Android'
                            )
                        )
                    ),
                    FilterExpression(
                        filter=Filter(
                            field_name='platform',
                            string_filter=Filter.StringFilter(
                                match_type=Filter.StringFilter.MatchType.EXACT,
                                value='iOS'
                            )
                        )
                    )
                ]
            )
        )
        
        request = RunReportRequest(
            property=f"properties/{property_id}",
            dimensions=[
                Dimension(name="month"),
                Dimension(name="year")
            ],
            metrics=[
                Metric(name="activeUsers"),
                Metric(name="sessions"),
                Metric(name="totalRevenue")
            ],
            date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
            dimension_filter=platform_filter,
            order_bys=[
                OrderBy(dimension={'dimension_name': 'year'}),
                OrderBy(dimension={'dimension_name': 'month'})
            ]
        )
        
        response = client.run_report(request=request)
        
        data = []
        for row in response.rows:
            month = row.dimension_values[0].value
            year = row.dimension_values[1].value
            
            data.append({
                'Mês': month,
                'Ano': year,
                'Usuários ativos': int(row.metric_values[0].value),
                'Sessões': int(row.metric_values[1].value),
                'Receita': float(row.metric_values[2].value)
            })
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            month_names = {
                '01': 'janeiro', '02': 'fevereiro', '03': 'março',
                '04': 'abril', '05': 'maio', '06': 'junho',
                '07': 'julho', '08': 'agosto', '09': 'setembro',
                '10': 'outubro', '11': 'novembro', '12': 'dezembro'
            }
            df['Mês'] = df['Mês'].map(month_names)
            
            logger.info(f"✓ Coletados {len(df)} meses")
            logger.info(f"  Total Usuários: {df['Usuários ativos'].sum():,.0f}".replace(',', '.'))
            logger.info(f"  Total Sessões: {df['Sessões'].sum():,.0f}".replace(',', '.'))
            logger.info(f"  Total Receita: R$ {df['Receita'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        return df
        
    except Exception as e:
        logger.error(f"Erro ao coletar Bemol App: {str(e)}", exc_info=True)
        return None

# ============================================================================
# FUNÇÕES DE COLETA - BEMOL FARMA
# ============================================================================

def fetch_bemol_farma_data(
    client: BetaAnalyticsDataClient,
    property_id: str,
    start_date: str,
    end_date: str
) -> Optional[pd.DataFrame]:
    """
    Coleta dados para Bemol Farma:
    - Sessões orgânicas
    - Usuários orgânicos
    - Taxa de engajamento (%)
    - Receita orgânica
    Filtro: sessionDefaultChannelGroup = 'Organic Search'
    """
    try:
        logger.info("🏥 Coletando dados: Bemol Farma - Tráfego Orgânico")
        
        # Filtro para tráfego orgânico
        organic_filter = FilterExpression(
            filter=Filter(
                field_name='sessionDefaultChannelGroup',
                string_filter=Filter.StringFilter(
                    match_type=Filter.StringFilter.MatchType.EXACT,
                    value='Organic Search'
                )
            )
        )
        
        request = RunReportRequest(
            property=f"properties/{property_id}",
            dimensions=[
                Dimension(name="month"),
                Dimension(name="year")
            ],
            metrics=[
                Metric(name="sessions"),
                Metric(name="activeUsers"),
                Metric(name="engagementRate"),
                Metric(name="totalRevenue")
            ],
            date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
            dimension_filter=organic_filter,
            order_bys=[
                OrderBy(dimension={'dimension_name': 'year'}),
                OrderBy(dimension={'dimension_name': 'month'})
            ]
        )
        
        response = client.run_report(request=request)
        
        data = []
        for row in response.rows:
            month = row.dimension_values[0].value
            year = row.dimension_values[1].value
            
            # Taxa de engajamento vem como decimal (0.75 = 75%)
            engagement_rate = float(row.metric_values[2].value) * 100
            
            data.append({
                'Mês': month,
                'Ano': year,
                'Sessões orgânicas': int(row.metric_values[0].value),
                'Usuários orgânicos': int(row.metric_values[1].value),
                'Taxa de engajamento (%)': round(engagement_rate, 2),
                'Receita orgânica': float(row.metric_values[3].value)
            })
        
        df = pd.DataFrame(data)
        
        if not df.empty:
            month_names = {
                '01': 'janeiro', '02': 'fevereiro', '03': 'março',
                '04': 'abril', '05': 'maio', '06': 'junho',
                '07': 'julho', '08': 'agosto', '09': 'setembro',
                '10': 'outubro', '11': 'novembro', '12': 'dezembro'
            }
            df['Mês'] = df['Mês'].map(month_names)
            
            logger.info(f"✓ Coletados {len(df)} meses")
            logger.info(f"  Total Sessões: {df['Sessões orgânicas'].sum():,.0f}".replace(',', '.'))
            logger.info(f"  Total Usuários: {df['Usuários orgânicos'].sum():,.0f}".replace(',', '.'))
            logger.info(f"  Engajamento médio: {df['Taxa de engajamento (%)'].mean():.2f}%")
            logger.info(f"  Total Receita: R$ {df['Receita orgânica'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        return df
        
    except Exception as e:
        logger.error(f"Erro ao coletar Bemol Farma: {str(e)}", exc_info=True)
        return None

# ============================================================================
# EXPORTAÇÃO EXCEL
# ============================================================================

def export_to_excel(
    df_bemol_web: pd.DataFrame,
    df_bemol_app: pd.DataFrame,
    df_bemol_farma: pd.DataFrame
) -> str:
    """Exporta todos os dados para Excel com 3 sheets"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'GA4_Bemol_Complete_{Config.ANALYSIS_YEAR}_{timestamp}.xlsx'
        filepath = os.path.join(Config.OUTPUT_DIR, filename)
        
        logger.info("📝 Gerando arquivo Excel...")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            if df_bemol_web is not None and not df_bemol_web.empty:
                df_bemol_web.to_excel(writer, sheet_name='Bemol (web)', index=False)
                logger.info("  ✓ Sheet: Bemol (web)")
            
            if df_bemol_app is not None and not df_bemol_app.empty:
                df_bemol_app.to_excel(writer, sheet_name='Bemol (App)', index=False)
                logger.info("  ✓ Sheet: Bemol (App)")
            
            if df_bemol_farma is not None and not df_bemol_farma.empty:
                df_bemol_farma.to_excel(writer, sheet_name='Bemol Farma', index=False)
                logger.info("  ✓ Sheet: Bemol Farma")
        
        # Formatação
        wb = load_workbook(filepath)
        
        # Cores e estilos
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formatar cabeçalho
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Ajustar larguras
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Formatar células de dados
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    col_letter = get_column_letter(col_idx)
                    
                    # Mês e Ano - alinhar à esquerda
                    if col_letter in ['A', 'B']:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Números inteiros (sessões, usuários)
                    elif cell.value and isinstance(cell.value, (int, float)) and 'Receita' not in str(ws[f'{col_letter}1'].value) and '%' not in str(ws[f'{col_letter}1'].value):
                        cell.number_format = '#,##0'
                    
                    # Receita (moeda)
                    elif cell.value and 'Receita' in str(ws[f'{col_letter}1'].value):
                        cell.number_format = 'R$ #,##0.00'
                    
                    # Percentual
                    elif cell.value and '%' in str(ws[f'{col_letter}1'].value):
                        cell.number_format = '0.00"%"'
            
            # Congelar primeira linha
            ws.freeze_panes = 'A2'
        
        wb.save(filepath)
        
        logger.info(f"✅ Arquivo Excel criado: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {str(e)}", exc_info=True)
        return ""

# ============================================================================
# MAIN
# ============================================================================

def main():
    """Função principal"""
    try:
        print("\n" + "="*80)
        print("GA4 COMPLETE DATA COLLECTOR - BEMOL GROUP")
        print("="*80)
        print(f"Período: {Config.ANALYSIS_START} a {Config.ANALYSIS_END}")
        print(f"Propriedades:")
        print(f"  • Ecommerce Bemol: {Config.PROPERTIES['ecommerce_bemol']}")
        print(f"  • Bemol Farma: {Config.PROPERTIES['bemol_farma']}")
        print("="*80 + "\n")
        
        # Autenticação
        client = authenticate_ga4()
        if not client:
            logger.error("❌ Falha na autenticação")
            return
        
        os.makedirs(Config.OUTPUT_DIR, exist_ok=True)
        
        # ====================================================================
        # COLETA DE DADOS
        # ====================================================================
        
        logger.info("\n" + "="*80)
        logger.info("INICIANDO COLETA DE DADOS")
        logger.info("="*80 + "\n")
        
        # 1. Bemol (web) - da propriedade Ecommerce Bemol
        df_bemol_web = fetch_bemol_web_data(
            client,
            Config.PROPERTIES['ecommerce_bemol'],
            Config.ANALYSIS_START,
            Config.ANALYSIS_END
        )
        
        print()  # Espaço visual
        
        # 2. Bemol (App) - da propriedade Ecommerce Bemol
        df_bemol_app = fetch_bemol_app_data(
            client,
            Config.PROPERTIES['ecommerce_bemol'],
            Config.ANALYSIS_START,
            Config.ANALYSIS_END
        )
        
        print()  # Espaço visual
        
        # 3. Bemol Farma - da propriedade Bemol Farma
        df_bemol_farma = fetch_bemol_farma_data(
            client,
            Config.PROPERTIES['bemol_farma'],
            Config.ANALYSIS_START,
            Config.ANALYSIS_END
        )
        
        # ====================================================================
        # EXPORTAÇÃO
        # ====================================================================
        
        logger.info("\n" + "="*80)
        logger.info("EXPORTANDO RESULTADOS")
        logger.info("="*80 + "\n")
        
        excel_path = export_to_excel(df_bemol_web, df_bemol_app, df_bemol_farma)
        
        # ====================================================================
        # RESUMO FINAL
        # ====================================================================
        
        print("\n" + "="*80)
        print("✅ COLETA CONCLUÍDA COM SUCESSO!")
        print("="*80)
        print(f"\n📁 Arquivo gerado: {excel_path}")
        print("\n📊 Resumo dos dados coletados:")
        print("-" * 80)
        
        if df_bemol_web is not None and not df_bemol_web.empty:
            print(f"\n🌐 BEMOL (WEB) - Ecommerce Bemol")
            print(f"   Meses: {len(df_bemol_web)}")
            print(f"   Sessões Totais: {df_bemol_web['Sessões totais (todos os canais)'].sum():,.0f}".replace(',', '.'))
            print(f"   Receita Total: R$ {df_bemol_web['Receita total (todos os canais)'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        if df_bemol_app is not None and not df_bemol_app.empty:
            print(f"\n📱 BEMOL (APP) - Ecommerce Bemol")
            print(f"   Meses: {len(df_bemol_app)}")
            print(f"   Usuários Ativos: {df_bemol_app['Usuários ativos'].sum():,.0f}".replace(',', '.'))
            print(f"   Sessões: {df_bemol_app['Sessões'].sum():,.0f}".replace(',', '.'))
            print(f"   Receita: R$ {df_bemol_app['Receita'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        if df_bemol_farma is not None and not df_bemol_farma.empty:
            print(f"\n🏥 BEMOL FARMA - Tráfego Orgânico")
            print(f"   Meses: {len(df_bemol_farma)}")
            print(f"   Sessões Orgânicas: {df_bemol_farma['Sessões orgânicas'].sum():,.0f}".replace(',', '.'))
            print(f"   Usuários Orgânicos: {df_bemol_farma['Usuários orgânicos'].sum():,.0f}".replace(',', '.'))
            print(f"   Taxa Engajamento Média: {df_bemol_farma['Taxa de engajamento (%)'].mean():.2f}%")
            print(f"   Receita Orgânica: R$ {df_bemol_farma['Receita orgânica'].sum():,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        
        print("\n" + "="*80 + "\n")
        
    except KeyboardInterrupt:
        print("\n⚠️  Execução interrompida pelo usuário")
    except Exception as e:
        logger.error(f"❌ Erro fatal: {str(e)}", exc_info=True)

if __name__ == "__main__":
    main()
